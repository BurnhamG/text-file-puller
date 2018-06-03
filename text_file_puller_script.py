#! python3
"""This will pull contract text files from the system.

Excel code based on examples from www.automatetheboringstuff.com/chapter12
"""
import glob
import os
import re
import time
import win32gui
from ctypes import windll
from datetime import datetime

import openpyxl
import psutil
import pyautogui


def list_email_groups(file_path):
    """Identify the groups of representatives to avoid certain files."""
    print('Identifying groups for emailing...')

    names_list = []

    with open(os.path.join(file_path, 'DataFiles', 'NonConList.txt'),
              encoding='utf-8') as f:
        names_list.append([line.rstrip() for line in f])

    with open(os.path.join(file_path, 'DataFiles', 'NoProcess.txt'),
              encoding='utf-8') as f:
        names_list.append([line.rstrip() for line in f])

    return names_list


def step_recognize(input_image):
    """Check for text on screen. If none matches, stop.

    In reserve in case this becomes necessary
    """
    ready_status = pyautogui.locateOnScreen(os.path.join(
        'DataFiles/Images', input_image),
        minSearchTime=.5)
    if not ready_status:
        print('Image not found.')
        return False


def get_keystrokes(file_path):
    """Get an array of the keystrokes used when pulling text files."""
    print('Obtaining keystrokes...')

    with open(os.path.join(file_path, 'DataFiles', 'Keystrokes.txt'),
              encoding='utf-8') as f:
        list_of_keystrokes = f.readlines()
        # Removes comments
        list_of_keystrokes = [x.split('|') for x in [x.replace(', ', '|').rstrip()
                                                     for x in list_of_keystrokes
                                                     if not x.startswith('#')]]

    return list_of_keystrokes


def get_window():
    """Set focus on the window and begin input."""
    for proc in psutil.process_iter():
        if re.search('.*mvbase.*', proc.name().lower()):
            hwnd = win32gui.FindWindow(None, proc.name())
            window_loc = win32gui.GetWindowRect(hwnd)
            return window_loc, hwnd
        else:
            print('No window detected.')
            input()
            raise SystemExit


def find_already_pulled(in_row, in_file_dict, avoid_list, source_sheet):
    """Mark files that have already been pulled with the pull time."""
    cellvalue = str(source_sheet.cell(row=in_row, column=5).value)
    contract_file = f'{cellvalue}.txt'

    if contract_file in in_file_dict:
        modified_date = in_file_dict[contract_file]
        source_sheet.cell(row=in_row, column=9).value = modified_date
    # Checks for text files to ignore
    if source_sheet.cell(row=in_row, column=8) in avoid_list:
        source_sheet.cell(row=in_row, column=9).value = 'IGNORED'


def get_contract_info(source_sheet):
    """Put contracts into a dictionary for future reference."""
    contract_info = {}
    for row in range(2, source_sheet.max_row + 1):
        count = 1
        contract_num = str(source_sheet['A' + str(row)].value)
        company_num = str(source_sheet['E' + str(row)].value)
        company_rep = str(source_sheet['G' + str(row)].value)
        contract_rep = str(source_sheet['H' + str(row)].value)

        # Ensure key for contract exists
        contract_info.setdefault(contract_num, {'Companies': [company_num],
                                                'CompanyRep': company_rep,
                                                'ContractRep': contract_rep,
                                                'CompanyCount': 1}
                                 )
        # Check for contract with multiple companies
        while source_sheet['A' + str(row + count)].value is None:
            contract_info[contract_num]['Companies'].append(str(source_sheet[
                                                                    'E' + str(row + count)]))
            count += 1
            contract_info[contract_num]['CompanyCount'] += 1
    return contract_info


def save_contract_files(all_contracts, select_contract, keys, non_con_reps, non=0):
    """Pull the files from the database."""
    count = 0
    current_time = datetime.now()
    txt_path = os.path.join('H:', os.sep, 'CONTXTFILES', select_contract)

    pyautogui.typewrite(str(current_time.month))
    pyautogui.typewrite('.01.')
    pyautogui.typewrite(str(current_time.year - 1))
    pyautogui.typewrite(['enter'])
    pyautogui.typewrite(['enter'])

    for company in all_contracts[select_contract]['Companies']:
        pyautogui.typewrite(company)
        pyautogui.typewrite(['enter'])
        pyautogui.typewrite(select_contract)
        pyautogui.typewrite(['enter'])
        count += 1

    pyautogui.typewrite(['enter'])

    if count > 1:
        pyautogui.typewrite(keys[4])
        pyautogui.typewrite(['enter'])

    if non != 0:
        pyautogui.typewrite(keys[6])
    else:
        pyautogui.typewrite(keys[5])
    pyautogui.typewrite(['enter'])
    pyautogui.typewrite(['enter'])
    pyautogui.typewrite(txt_path)

    if non == 1:
        pyautogui.typewrite('NON')

    pyautogui.typewrite(['enter'] * 2)
    time.sleep(10)

    if all_contracts[select_contract]['CompanyRep'] in non_con_reps and non == 0:
        save_contract_files(all_contracts, select_contract, keys, non_con_reps, non=1)


def menu_setup(win_x, win_width, win_y, win_height, in_keystrokes):
    """Prepare the database screen for the input."""
    pyautogui.click(win_x + (win_width / 2), win_y + (win_height / 2))
    pyautogui.typewrite(['enter'] * 5)
    for key in in_keystrokes:
        pyautogui.typewrite(key)
        pyautogui.typewrite(['enter'])


def choose_text_file_directory(file_path):
    text_file_path = os.path.join(file_path, 'Contract Renewal Text Files')
    create_dir = None
    print('Where would you like the text files stored?')
    print('Please enter the full path.')
    text_file_path_selection = input('Default is ' + text_file_path + ': ')
    if text_file_path_selection != '':
        text_file_path = text_file_path_selection
    if os.path.exists(text_file_path):
        os.chdir(text_file_path)
    else:
        print('This directory does not exist.')
        while not create_dir:
            create_dir = input("Would you like to create it? ([y]es/[n]o) ")
        if create_dir.lower() == 'y':
            try:
                os.makedirs(text_file_path)
            except OSError as oe:
                print('The system encountered an error:')
                print(oe)
                raise SystemExit
        else:
            print('Goodbye!')
            raise SystemExit


def get_source_spreadsheet():
    search_here = None
    list_of_excel_files = None
    print(f'The current directory is {os.getcwd()}')
    while not search_here:
        search_here = input('Is this where you would like to search for the source spreadsheet ([y]es/[n]o/[e]xit)? ')
        if search_here.lower() == 'e':
            print('Goodbye!')
            raise SystemExit
        elif search_here.lower() == 'n':
            search_here = input(
                'Where would you like to search for the source spreadsheet? Please enter the full path: ')
        elif search_here.lower() != 'y':
            print('That is not a valid response, please try again.')
            search_here = None
        if os.path.exists(search_here):
            os.chdir(search_here)
            list_of_excel_files = glob.glob('*.xlsx')
        else:
            print('That is not a valid directory, please try again.')
            search_here = None

    if not list_of_excel_files:
        print('Excel files not found.')
        raise SystemExit
    for file in list_of_excel_files:
        use_file = input('Is ' + file + ' the source spreadsheet ([y]es/[n]o)? ')
        if use_file.lower() == 'y':
            source_book = openpyxl.load_workbook(file)
            return source_book[source_book.sheetnames[0]]


def get_file_modification_times():
    list_of_files = glob.glob('*.txt')
    mod_times = {}

    # Modification time of the files
    for f in list_of_files:
        mod_times[f] = datetime.fromtimestamp(time.mktime(time.localtime(os.stat(f).st_mtime)))

    return mod_times


def pull_contract_files():
    this_file_path = os.path.abspath(__file__)
    window_coords, hwnd = get_window()
    win_x = window_coords[0]
    win_y = window_coords[1]
    win_width = window_coords[2] - win_x
    win_height = window_coords[3] - win_y
    set_win_pos = windll.user32.SetWindowPos
    contracts_to_pull = -1
    start_contract = None
    end_contract = None

    choose_text_file_directory(this_file_path)
    get_file_modification_times()
    keystrokes = get_keystrokes(this_file_path)
    source_sheet = get_source_spreadsheet()
    file_modifications = get_file_modification_times()

    all_names = list_email_groups(this_file_path)
    non_con_reps = all_names[0]
    list_of_contracts_to_avoid = all_names[1]

    # Checks for files that have already been pulled
    for i in range(2, source_sheet.max_row + 1):  # Skips header row
        find_already_pulled(i, file_modifications, list_of_contracts_to_avoid, source_sheet)

    while contracts_to_pull is not -1:
        # Ask the user for the first contract they would like to pull, or give the user
        # the option of inputting a list of contracts they would like to pull
        # (i.e. ones that came out blank) or leaving the input empty to pull all files.
        print('What contracts would you like to pull?')
        print('If entering a range, separate the numbers with a hyphen, or')
        print('use commas to separate individual contracts.')
        print('Leave this empty to pull all text files.')
        print('Type "end" to exit.')
        contracts_to_pull = input()

        if contracts_to_pull.strip() == 'end':
            print("Goodbye!")
            raise SystemExit
        elif re.search(r'[^0-9-,]', contracts_to_pull, re.I):
            print("You entered an invalid character.")
            contracts_to_pull = -1
        elif ',' in contracts_to_pull:
            contracts_to_pull = contracts_to_pull.split(',')
        elif contracts_to_pull.strip() != '':
            # Separate the beginning and end of the range.
            start_contract = int(contracts_to_pull.split('-')[0])
            end_contract = int(contracts_to_pull.split('-')[1])

    all_contracts = get_contract_info(source_sheet)
    # Setting current time = datetime.now().strftime('%m-%d-%Y %H:%M:%S')

    # Go through, pulling text files and saving them under the contract name.
    # Make sure this also pulls the non-contract items if that applies.

    # The -1 locks the window on top.
    set_win_pos(hwnd, -1, win_x, win_y, 0, 0, 0x0001)

    # Start interacting with the database
    menu_setup(win_x, win_width, win_y, win_height, keystrokes)
    if start_contract:
        for contract in range(start_contract, end_contract + 1):
            if contract in all_contracts:
                save_contract_files(all_contracts, contract, keystrokes, non_con_reps)
    elif isinstance(contracts_to_pull, list):
        for contract in contracts_to_pull:
            if contract in all_contracts:
                save_contract_files(all_contracts, contract, keystrokes, non_con_reps)

    for contract in sorted(all_contracts):
        save_contract_files(all_contracts, contract, keystrokes, non_con_reps)
    set_win_pos(hwnd, 1, win_x, win_y, 0, 0, 0x0001)

    print('All files pulled!')
    input()


if __name__ == '__main__':
    pull_contract_files()
